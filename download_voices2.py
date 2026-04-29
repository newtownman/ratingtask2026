import requests
import re
import time
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_KEY   = "your_api_key_here"
AUDIO_DIR = Path("audio_samples_nonsouthern")
AUDIO_DIR.mkdir(exist_ok=True)

ACCENT_COLORS = {
    "boston":       "D5F5E3",
    "chicago":      "FDEBD0",
    "new york":     "F9EBEA",
    "us midwest":   "EAF0FB",
    "us northeast": "FEF9E7",
}

def fetch_voices(accent, limit):
    voices = []
    page = 0
    while True:
        res = requests.get(
            "https://api.elevenlabs.io/v1/shared-voices",
            headers={"xi-api-key": API_KEY},
            params={"language": "en", "accent": accent, "page_size": 100, "page": page},
            timeout=15,
        )
        data = res.json()
        batch = next((v for v in data.values() if isinstance(v, list) and len(v) > 0), [])
        voices.extend(batch)
        if not data.get("has_more"): break
        page += 1
        time.sleep(0.25)
    voices.sort(key=lambda v: v.get("usage_character_count_1y") or 0, reverse=True)
    return voices[:limit]

def build_excel(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Additional Voices"

    headers = [
        "accent", "rank", "filename",
        "voice_id", "name", "gender", "age", "category",
        "usage_1y", "usage_7d", "cloned_by",
        "free_users", "notice_period_days",
        "preview_url", "description",
    ]

    hdr_fill = PatternFill("solid", fgColor="2C3E50")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, row in enumerate(rows, 2):
        fill = PatternFill("solid", fgColor=ACCENT_COLORS.get(row["accent"], "FFFFFF"))
        values = [
            row["accent"], row["rank"], row["filename"],
            row["voice_id"], row["name"], row["gender"], row["age"], row["category"],
            row["usage_1y"], row["usage_7d"], row["cloned_by"],
            row["free_users"], row["notice_period"],
            row["preview_url"], row["description"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == len(headers)))

    col_widths = [14, 6, 38, 26, 22, 8, 8, 14, 14, 12, 10, 10, 14, 50, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"\n✅ 엑셀 저장: {out_path}")

# ── 메인 ──
all_rows = []

for accent, slug in [("new york", "new_york"), ("us midwest", "us_midwest")]:
    voices = fetch_voices(accent, 19)
    for i, voice in enumerate(voices[12:19], 13):
        vid      = voice.get("voice_id", "unknown")
        filename = f"{slug}__{i:03d}__{vid}.mp3"
        url      = voice.get("preview_url", "")
        if url:
            r = requests.get(url, timeout=20)
            if r.status_code == 200:
                (AUDIO_DIR / filename).write_bytes(r.content)
                print(f"✓ [{i:03d}] {voice.get('name','?')} → {filename}")

        all_rows.append({
            "accent":        accent,
            "rank":          i,
            "filename":      filename,
            "voice_id":      vid,
            "name":          voice.get("name", ""),
            "gender":        voice.get("gender", ""),
            "age":           voice.get("age", ""),
            "category":      voice.get("category", ""),
            "usage_1y":      voice.get("usage_character_count_1y"),
            "usage_7d":      voice.get("usage_character_count_7d"),
            "cloned_by":     voice.get("cloned_by_count"),
            "free_users":    voice.get("free_users_allowed"),
            "notice_period": voice.get("notice_period"),
            "preview_url":   url,
            "description":   voice.get("description", ""),
        })

print(f"\n총 {len(all_rows)}개 추가")
build_excel(all_rows, Path("additional_voices_report.xlsx"))
