"""
ElevenLabs Focus Condition Check — v2
- 각 보이스는 target word 하나에 대해 2개 파일 생성
- voice 1 → word 1, voice 2 → word 2, ... (word 수 초과 시 순환)
- 총 파일 수: 122 × 2 = 244
"""

import os
import time
import requests
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

# ── 설정 ──────────────────────────────────────────────
API_KEY       = "ffd6c21d6ab10c9d0a66aa382b2e63b79c4eaed93ab794b59950b8330eaddfb3"   # ← 여기에 입력
OUTPUT_DIR    = Path("focus_check")
MODEL_ID      = "eleven_multilingual_v2"
REQUEST_DELAY = 0.5   # 초 (rate limit 방지)

RED  = 'FFFF0000'
GRAY = 'FF808080'


# ── 1. Target words ───────────────────────────────────

def get_mono_words(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb['Mono_Freq']
    categories, current_cat, current_words = {}, None, []

    for row in ws.iter_rows(values_only=False):
        vals   = [c.value for c in row]
        colors = [
            c.font.color.rgb if (c.font and c.font.color and c.font.color.type == 'rgb')
            else 'none' for c in row
        ]
        if vals[0] and isinstance(vals[0], str) and '▶' in str(vals[0]):
            if current_cat and current_words:
                categories[current_cat] = current_words
            current_cat, current_words = vals[0], []
            continue
        if vals[1] and isinstance(vals[1], str) and vals[1] != '단어':
            if colors[1] not in [RED, GRAY]:
                current_words.append(vals[1].strip())

    if current_cat and current_words:
        categories[current_cat] = current_words

    return [w for words in categories.values() for w in words[:7]]  # 35개


def get_pinpen_words():
    nasal    = ['when','win','men','min','ten','tin','ben','bin',
                'pen','pin','jen','jin','ken','kin']
    nonnasal = ['get','git','dead','did','let','lit','set','sit',
                'bet','bit','head','hid','red','rid']
    return nasal + nonnasal  # 28개


def build_word_list(xlsx_path):
    words = get_mono_words(xlsx_path) + get_pinpen_words()  # 63개
    print(f"[INFO] Target words: {len(words)}")
    return words


# ── 2. Voice list ─────────────────────────────────────

def get_voices(voice_list_path):
    df = pd.read_excel(voice_list_path)
    voices = []
    for _, row in df.iterrows():
        region   = str(row['accent']).strip().lower().replace(' ', '_')
        name     = str(row['name']).strip().replace(' ', '_').replace("'", "")
        voice_id = str(row['voice_id']).strip()
        no       = int(row['no'])
        voices.append((no, region, name, voice_id))
    voices.sort(key=lambda x: x[0])
    print(f"[INFO] Total voices: {len(voices)}")
    return voices


# ── 3. TTS ────────────────────────────────────────────

def make_sentences(word):
    unfocused = f"I 'really' said {word} again."
    focused   = f'I said "{word}", again.'
    return unfocused, focused


def generate_audio(text, voice_id):
    url = f"https://api.elevenlabs.io/v1/text-to-speech/{voice_id}"
    headers = {
        "Accept": "audio/mpeg",
        "Content-Type": "application/json",
        "xi-api-key": API_KEY,
    }
    payload = {
        "text": text,
        "model_id": MODEL_ID,
        "voice_settings": {"stability": 0.5, "similarity_boost": 0.75}
    }
    resp = requests.post(url, json=payload, headers=headers)
    if resp.status_code == 200:
        return resp.content
    print(f"  [ERROR] {resp.status_code}: {resp.text[:120]}")
    return None


# ── 4. Main ───────────────────────────────────────────

def main():
    freq_path  = "frequency_checked_v5.xlsx"
    voice_path = "voice_list_121.xlsx"

    if API_KEY == "YOUR_ELEVENLABS_API_KEY":
        print("[ERROR] API 키를 입력해주세요")
        return

    OUTPUT_DIR.mkdir(exist_ok=True)

    words  = build_word_list(freq_path)
    voices = get_voices(voice_path)

    total = len(voices) * 2
    done = errors = 0

    print(f"[INFO] 총 파일 수: {total}  ({len(voices)} voices × 2)\n")

    for no, region, name, voice_id in voices:
        word = words[(no - 1) % len(words)]   # no 1부터 순서대로 단어 배정
        sent1, sent2 = make_sentences(word)

        for idx, sentence in enumerate([sent1, sent2], start=1):
            filename = f"{region}__{name}__{word}__{idx}.mp3"
            filepath = OUTPUT_DIR / filename

            if filepath.exists():
                print(f"  [SKIP] {filename}")
                done += 1
                continue

            audio = generate_audio(sentence, voice_id)
            if audio:
                filepath.write_bytes(audio)
                print(f"  [OK]   {filename}")
                done += 1
            else:
                print(f"  [FAIL] {filename}")
                errors += 1

            time.sleep(REQUEST_DELAY)

    print(f"\n[완료] 성공: {done} / 오류: {errors} / 전체: {total}")


if __name__ == "__main__":
    main()
