"""
ElevenLabs 121 Voice Collector
================================
선정 기준:
  us southern  → 전체 (약 61개)
  boston       → 전체 (약 7개)
  chicago      → 전체 (약 3개)
  us northeast → 전체 (약 12개)
  new york     → usage 상위 19개
  us midwest   → usage 상위 19개
  ──────────────────────────────
  합계 목표: 121개

출력:
  - audio/ 폴더에 mp3 다운로드
  - voice_list_121.xlsx 생성
"""

import requests
import re
import time
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────
#  CONFIG
# ────────────────────────────────────────
API_KEY   = "ffd6c21d6ab10c9d0a66aa382b2e63b79c4eaed93ab794b59950b8330eaddfb3"
AUDIO_DIR = Path("audio")
AUDIO_DIR.mkdir(exist_ok=True)

# None = 전체, 'fill' = 나머지 슬롯 자동 배분
NON_SOUTHERN_TOTAL = 60  # non-southern 목표 총합

TARGET_ACCENTS = {
    "us southern":  None,    # 전체
    "boston":       None,    # 전체
    "chicago":      None,    # 전체
    "us northeast": None,    # 전체
    "new york":     "fill",  # 자동 계산
    "us midwest":   "fill",  # 자동 계산
}

ACCENT_COLORS = {
    "us southern":  "D6EAF8",
    "boston":       "D5F5E3",
    "chicago":      "FDEBD0",
    "us northeast": "FEF9E7",
    "new york":     "F9EBEA",
    "us midwest":   "EAF0FB",
}

# ────────────────────────────────────────
#  FETCH
# ────────────────────────────────────────
def fetch_voices(accent, limit):
    all_voices = []
    page = 0
    print(f"\n[{accent}] 수집 중...", end="", flush=True)
    while True:
        res = requests.get(
            "https://api.elevenlabs.io/v1/shared-voices",
            headers={"xi-api-key": API_KEY},
            params={"language": "en", "accent": accent, "page_size": 100, "page": page},
            timeout=15,
        )
        if res.status_code != 200:
            print(f"\n  ⚠ HTTP {res.status_code}")
            break
        data = res.json()
        batch = next((v for v in data.values() if isinstance(v, list) and len(v) > 0), [])
        all_voices.extend(batch)
        print(f" {len(all_voices)}개", end="", flush=True)
        if not data.get("has_more"):
            break
        page += 1
        time.sleep(0.25)

    all_voices.sort(key=lambda v: v.get("usage_character_count_1y") or 0, reverse=True)
    result = all_voices[:limit] if limit else all_voices
    print(f" → 선정 {len(result)}개")
    return result

# ────────────────────────────────────────
#  DOWNLOAD
# ────────────────────────────────────────
def safe_slug(text):
    return re.sub(r"[^\w]", "_", text)

def safe_name(name):
    name = re.sub(r'[\\/:*?"<>|]', '', name)
    return name.strip().replace(' ', '_')[:40]

def download_audio(url, dest):
    if dest.exists():
        return True
    try:
        r = requests.get(url, timeout=20)
        if r.status_code == 200:
            dest.write_bytes(r.content)
            return True
        return False
    except:
        return False

# ────────────────────────────────────────
#  EXCEL
# ────────────────────────────────────────
def build_excel(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Voice List"

    headers = [
        "no", "accent", "rank_in_accent", "filename",
        "voice_id", "name", "gender", "age", "category",
        "usage_1y", "usage_7d", "cloned_by",
        "free_users_allowed", "notice_period_days",
        "preview_url", "description",
    ]

    hdr_fill = PatternFill("solid", fgColor="2C3E50")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, row in enumerate(rows, 1):
        excel_row = i + 1
        fill = PatternFill("solid", fgColor=ACCENT_COLORS.get(row["accent"], "FFFFFF"))
        values = [
            i,
            row["accent"], row["rank"], row["filename"],
            row["voice_id"], row["name"], row["gender"], row["age"], row["category"],
            row["usage_1y"], row["usage_7d"], row["cloned_by"],
            row["free_users"], row["notice_period"],
            row["preview_url"], row["description"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == len(headers)))

    col_widths = [5, 14, 14, 42, 26, 28, 8, 8, 14, 14, 12, 10, 16, 14, 50, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Summary 시트
    ws2 = wb.create_sheet("Summary")
    ws2.append(["accent", "선정 수", "총 usage_1y"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for accent in TARGET_ACCENTS:
        accent_rows = [r for r in rows if r["accent"] == accent]
        total_usage = sum(r["usage_1y"] or 0 for r in accent_rows)
        ws2.append([accent, len(accent_rows), total_usage])
    ws2.append(["합계", len(rows), sum(r["usage_1y"] or 0 for r in rows)])
    for col in ["A", "B", "C"]:
        ws2.column_dimensions[col].width = 18

    wb.save(out_path)
    print(f"✅ 엑셀 저장: {out_path}")

# ────────────────────────────────────────
#  MAIN
# ────────────────────────────────────────
def main():
    all_rows = []

    # 1단계: fill이 아닌 accent 먼저 수집해서 non-southern 고정 수 파악
    fixed_accents = {k: v for k, v in TARGET_ACCENTS.items() if v != "fill"}
    fill_accents  = [k for k, v in TARGET_ACCENTS.items() if v == "fill"]

    collected = {}
    for accent, limit in fixed_accents.items():
        collected[accent] = fetch_voices(accent, limit)

    # non-southern 고정분 합산
    fixed_non_southern = sum(
        len(collected[a]) for a in fixed_accents
        if a != "us southern"
    )
    remaining   = NON_SOUTHERN_TOTAL - fixed_non_southern
    fill_each   = remaining // len(fill_accents) if fill_accents else 0
    print(f"\n고정 non-southern: {fixed_non_southern}개 → fill accent당 {fill_each}개씩")

    for accent in fill_accents:
        collected[accent] = fetch_voices(accent, fill_each)

    # 2단계: 다운로드 + 엑셀 rows 구성 (accent 순서 유지)
    for accent in TARGET_ACCENTS:
        voices      = collected[accent]
        accent_slug = safe_slug(accent)

        for rank, voice in enumerate(voices, 1):
            vid      = voice.get("voice_id", "unknown")
            vname    = safe_name(voice.get("name", "unknown"))
            filename = f"{accent_slug}__{rank:03d}__{vname}.mp3"
            url      = voice.get("preview_url", "")

            if url:
                ok   = download_audio(url, AUDIO_DIR / filename)
                mark = "✓" if ok else "✗"
            else:
                mark = "⚠"

            print(f"  {mark} [{rank:03d}] {voice.get('name','?'):30s} → {filename}")

            all_rows.append({
                "accent":        accent,
                "rank":          rank,
                "filename":      filename,
                "voice_id":      vid,
                "name":          voice.get("name", ""),
                "gender":        voice.get("gender", ""),
                "age":           voice.get("age", ""),
                "category":      voice.get("category", ""),
                "usage_1y":      voice.get("usage_character_count_1y"),
                "usage_7d":      voice.get("usage_character_count_7d"),
                "cloned_by":     voice.get("cloned_by_count"),
                "free_users":    voice.get("free_users_allowed"),
                "notice_period": voice.get("notice_period"),
                "preview_url":   url,
                "description":   voice.get("description", ""),
            })

    print(f"\n총 {len(all_rows)}개 선정 완료")
    build_excel(all_rows, Path("voice_list_121.xlsx"))
    print(f"오디오: {AUDIO_DIR}/")

if __name__ == "__main__":
    main()
