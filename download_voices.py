"""
ElevenLabs Voice Sample Downloader
====================================
- us southern: 전체 수집
- boston / chicago / new york / us midwest / us northeast: 인기 상위 12개씩
- preview mp3 다운로드 → audio_samples/ 폴더
- 메타데이터 + 사용량 → voice_report.xlsx
"""

import requests
import os
import time
from pathlib import Path
import re

# ── pip install openpyxl 필요 ──
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────
#  CONFIG
# ────────────────────────────────────────
API_KEY   = "ffd6c21d6ab10c9d0a66aa382b2e63b79c4eaed93ab794b59950b8330eaddfb3"
AUDIO_DIR = Path("audio_samples")
AUDIO_DIR.mkdir(exist_ok=True)

# accent → 가져올 최대 수 (None = 전체)
TARGET_ACCENTS = {
    "us southern":  None,
    "boston":       12,
    "chicago":      12,
    "new york":     12,
    "us midwest":   12,
    "us northeast": 12,
}

# ────────────────────────────────────────
#  FETCH VOICES
# ────────────────────────────────────────
def fetch_voices(accent: str, limit: int | None) -> list[dict]:
    """accent 기준으로 전체 페이지 수집 후 사용량순 정렬."""
    all_voices = []
    page = 0

    print(f"\n[{accent}] 수집 중...", end="")
    while True:
        res = requests.get(
            "https://api.elevenlabs.io/v1/shared-voices",
            headers={"xi-api-key": API_KEY},
            params={
                "language":   "en",
                "accent":     accent,
                "page_size":  100,
                "page":       page,
            },
            timeout=15,
        )
        if res.status_code != 200:
            print(f"\n  ⚠ HTTP {res.status_code}: {res.text[:200]}")
            break

        data = res.json()
        # 리스트 키 자동 탐색 (voices / shared_voices 등)
        voices = next(
            (v for v in data.values() if isinstance(v, list) and len(v) > 0),
            []
        )
        all_voices.extend(voices)
        print(f" {len(all_voices)}개", end="", flush=True)

        if not data.get("has_more"):
            break
        page += 1
        time.sleep(0.25)   # rate limit 여유

    print(f" → 총 {len(all_voices)}개 수집")

    # 연간 사용량 내림차순 정렬
    all_voices.sort(
        key=lambda v: v.get("usage_character_count_1y") or 0,
        reverse=True,
    )

    return all_voices[:limit] if limit else all_voices


# ────────────────────────────────────────
#  DOWNLOAD AUDIO
# ────────────────────────────────────────
def safe_filename(accent: str, voice: dict, idx: int) -> str:
    """accent__순번__voiceID.mp3 형태로 파일명 생성."""
    accent_slug = re.sub(r"[^\w]", "_", accent)
    voice_id    = voice.get("voice_id", "unknown")
    return f"{accent_slug}__{idx:03d}__{voice_id}.mp3"


def download_audio(url: str, dest: Path) -> bool:
    if dest.exists():
        return True   # 이미 있으면 스킵
    try:
        r = requests.get(url, timeout=20)
        if r.status_code == 200:
            dest.write_bytes(r.content)
            return True
        print(f"  ⚠ 다운로드 실패 ({r.status_code}): {dest.name}")
        return False
    except Exception as e:
        print(f"  ⚠ 오류: {e}")
        return False


# ────────────────────────────────────────
#  EXCEL REPORT
# ────────────────────────────────────────
ACCENT_COLORS = {
    "us southern":  "D6EAF8",
    "boston":       "D5F5E3",
    "chicago":      "FDEBD0",
    "new york":     "F9EBEA",
    "us midwest":   "EAF0FB",
    "us northeast": "FEF9E7",
}

def build_excel(all_rows: list[dict], out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Voice Report"

    headers = [
        "accent", "rank", "filename",
        "voice_id", "name", "gender", "age", "category",
        "usage_1y", "usage_7d", "cloned_by",
        "free_users", "notice_period_days",
        "preview_url", "description",
    ]

    # Header row
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2C3E50")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align

    ws.row_dimensions[1].height = 28

    # Data rows
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_data in all_rows:
        accent   = row_data["accent"]
        row_idx  = row_data["row_idx"]
        fill_hex = ACCENT_COLORS.get(accent, "FFFFFF")
        fill     = PatternFill("solid", fgColor=fill_hex)

        values = [
            accent,
            row_data["rank"],
            row_data["filename"],
            row_data["voice_id"],
            row_data["name"],
            row_data["gender"],
            row_data["age"],
            row_data["category"],
            row_data["usage_1y"],
            row_data["usage_7d"],
            row_data["cloned_by"],
            row_data["free_users"],
            row_data["notice_period"],
            row_data["preview_url"],
            row_data["description"],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill   = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == len(headers)))

    # Column widths
    col_widths = [14, 6, 38, 26, 22, 8, 8, 14, 14, 12, 10, 10, 14, 50, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["accent", "count", "총 usage_1y"])
    summary_font = Font(bold=True)
    for cell in ws2[1]:
        cell.font = summary_font

    for accent, count, total_usage in sorted(
        set((r["accent"], None, None) for r in all_rows),
        key=lambda x: x[0]
    ):
        rows_for = [r for r in all_rows if r["accent"] == accent]
        total    = sum(r["usage_1y"] or 0 for r in rows_for)
        ws2.append([accent, len(rows_for), total])

    for col in ["A", "B", "C"]:
        ws2.column_dimensions[col].width = 18

    wb.save(out_path)
    print(f"\n✅ 엑셀 저장: {out_path}")


# ────────────────────────────────────────
#  MAIN
# ────────────────────────────────────────
def main():
    all_rows   = []
    excel_row  = 2   # row 1 = header

    for accent, limit in TARGET_ACCENTS.items():
        voices = fetch_voices(accent, limit)

        for rank, voice in enumerate(voices, 1):
            filename    = safe_filename(accent, voice, rank)
            preview_url = voice.get("preview_url", "")

            # 오디오 다운로드
            if preview_url:
                dest = AUDIO_DIR / filename
                ok   = download_audio(preview_url, dest)
                status = "✓" if ok else "✗"
                print(f"  {status} [{rank:03d}] {voice.get('name', '?')} → {filename}")
            else:
                print(f"  ⚠ preview_url 없음: {voice.get('name', '?')}")

            all_rows.append({
                "accent":       accent,
                "rank":         rank,
                "filename":     filename,
                "voice_id":     voice.get("voice_id", ""),
                "name":         voice.get("name", ""),
                "gender":       voice.get("gender", ""),
                "age":          voice.get("age", ""),
                "category":     voice.get("category", ""),
                "usage_1y":     voice.get("usage_character_count_1y"),
                "usage_7d":     voice.get("usage_character_count_7d"),
                "cloned_by":    voice.get("cloned_by_count"),
                "free_users":   voice.get("free_users_allowed"),
                "notice_period": voice.get("notice_period"),
                "preview_url":  preview_url,
                "description":  voice.get("description", ""),
                "row_idx":      excel_row,
            })
            excel_row += 1

    print(f"\n총 {len(all_rows)}개 보이스 수집 완료")
    build_excel(all_rows, Path("voice_report.xlsx"))
    print(f"오디오 파일: {AUDIO_DIR}/")


if __name__ == "__main__":
    main()
